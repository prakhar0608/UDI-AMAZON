"""Export CSV and XLSX with formatting."""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


def sanitize_filename_part(name: str) -> str:
    s = re.sub(r"[^\w\-.]+", "_", name.strip(), flags=re.UNICODE)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "account"


def _apply_excel_formatting(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str = "Summary"):
    """Apply styles, number formats, and colors to the worksheet."""
    workbook = writer.book
    if sheet_name not in workbook.sheetnames:
        return
    worksheet = workbook[sheet_name]

    # Header styling
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)

    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column-specific formatting
    for col_num, col_name in enumerate(df.columns, 1):
        # Determine format based on column name
        num_format = None
        fill = None

        col_lower = col_name.lower()
        if "_id" in col_lower:
            # IDs should be strings/integers, no decimals
            num_format = "0"
        elif "spend" in col_lower or "sales" in col_lower:
            # "replace decimal" - display as integers
            num_format = "#,##0"
            fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        elif "roas" in col_lower or "cpc" in col_lower:
            # ROAS and CPC: 2 decimal points
            num_format = "0.00"
            fill = PatternFill(start_color="EBF5EB", end_color="EBF5EB", fill_type="solid")
        elif "acos" in col_lower or "ctr" in col_lower:
            # ACoS and CTR: 2 decimal points with percentage sign
            num_format = '0.00"%"'
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        if num_format or fill:
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                if num_format:
                    cell.number_format = num_format
                if fill:
                    cell.fill = fill

    # Auto-adjust column width
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)


def export_dataframe(
    summary_df: pd.DataFrame,
    performance_df: pd.DataFrame,
    out_dir: Path,
    *,
    stem: str,
) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / f"{stem}.csv"
    xlsx_path = out_dir / f"{stem}.xlsx"

    # Save performance format to CSV (it's the main data)
    performance_df.to_csv(csv_path, index=False)

    # Save to Excel with two sheets: Summary and Campaign Performance
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        _apply_excel_formatting(writer, summary_df, "Summary")
        
        performance_df.to_excel(writer, index=False, sheet_name="Campaign Performance")
        _apply_excel_formatting(writer, performance_df, "Campaign Performance")

    return csv_path, xlsx_path


def dataframe_to_csv_xlsx_bytes(wide_df: pd.DataFrame, long_df: pd.DataFrame | None = None) -> tuple[bytes, bytes]:
    """Convert DataFrames to CSV/XLSX bytes without writing files."""
    csv_bytes = wide_df.to_csv(index=False).encode("utf-8")

    xlsx_io = BytesIO()
    with pd.ExcelWriter(xlsx_io, engine="openpyxl") as writer:
        wide_df.to_excel(writer, index=False, sheet_name="Summary")
        _apply_excel_formatting(writer, wide_df, "Summary")
        
        if long_df is not None and not long_df.empty:
            long_df.to_excel(writer, index=False, sheet_name="Daily Performance")
            _apply_excel_formatting(writer, long_df, "Daily Performance")
    
    xlsx_bytes = xlsx_io.getvalue()
    xlsx_io.close()
    return csv_bytes, xlsx_bytes

